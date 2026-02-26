# PDF to Excel Conversion
from PyPDF2 import PdfReader
import pdfplumber
from pathlib import Path
from typing import List, Optional
import io


class PDFToExcelConverter:
    """PDF转Excel转换器"""

    def __init__(self, pdf_path: Path, output_path: Path):
        self.pdf_path = pdf_path
        self.output_path = output_path

    def extract_tables(self) -> dict:
        """从PDF中提取表格数据"""
        try:
            tables_data = []

            # 使用pdfplumber提取表格
            with pdfplumber.open(str(self.pdf_path)) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    tables = page.extract_tables()

                    if tables:
                        for table in tables:
                            # 清理表格数据
                            clean_table = self._clean_table(table)
                            if clean_table:
                                tables_data.append({
                                    "page": page_num + 1,
                                    "data": clean_table,
                                    "rows": len(clean_table),
                                    "cols": len(clean_table[0]) if clean_table else 0,
                                })

            return {
                "success": True,
                "tables": tables_data,
                "table_count": len(tables_data),
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
            }

    def extract_text_to_excel(self) -> dict:
        """将PDF文本提取为Excel格式"""
        try:
            reader = PdfReader(str(self.pdf_path))
            text_data = []

            for page_num, page in enumerate(reader.pages):
                text = page.extract_text()
                if text.strip():
                    text_data.append({
                        "page": page_num + 1,
                        "content": text,
                    })

            return {
                "success": True,
                "text_data": text_data,
                "page_count": len(text_data),
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
            }

    def convert(self, method: str = "auto") -> dict:
        """将PDF转换为Excel格式

        Args:
            method: 转换方法
                "auto" - 自动检测（优先提取表格）
                "table" - 只提取表格
                "text" - 将文本提取为Excel
        """
        try:
            if method == "auto":
                # 首先尝试提取表格
                result = self.extract_tables()
                if not result.get("success") or result.get("table_count", 0) == 0:
                    # 如果没有表格，则提取文本
                    result = self.extract_text_to_excel()
            elif method == "table":
                result = self.extract_tables()
            elif method == "text":
                result = self.extract_text_to_excel()
            else:
                return {
                    "success": False,
                    "error": f"不支持的转换方法: {method}",
                }

            if not result.get("success"):
                return result

            # 使用openpyxl创建Excel文件
            return self._create_excel_file(result, method)

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
            }

    def _clean_table(self, table: list) -> List[List[str]]:
        """清理表格数据"""
        if not table:
            return []

        clean_table = []
        for row in table:
            clean_row = []
            for cell in row:
                # 清理单元格内容
                if cell is None:
                    clean_row.append("")
                else:
                    text = str(cell).strip()
                    # 合并多行文本
                    text = text.replace("\n", " ")
                    clean_row.append(text)
            if clean_row:
                clean_table.append(clean_row)

        return clean_table

    def _create_excel_file(self, extract_result: dict, method: str) -> dict:
        """创建Excel文件"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # 定义样式
        header_font = Font(name='Arial', size=12, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        if method == "table" and extract_result.get("tables"):
            # 写入表格数据
            table_data = extract_result["tables"]
            if table_data:
                # 使用第一个表格作为模板
                first_table = table_data[0]["data"]
                headers = [f"列{i+1}" for i in range(len(first_table[0]))]

                # 写入表头
                for col, header in enumerate(headers):
                    cell = ws.cell(row=1, column=col + 1, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = header_border

                # 写入数据
                row_offset = 2
                for table in table_data:
                    for row_idx, row_data in enumerate(table["data"]):
                        for col_idx, cell_value in enumerate(row_data):
                            cell = ws.cell(
                                row=row_offset + row_idx,
                                column=col_idx + 1,
                                value=cell_value
                            )
                ws[f"A{row_offset+len(table_data)}"] = f"（来自第{table['page']}页）"

        elif method == "text" and extract_result.get("text_data"):
            # 写入文本数据
            ws.append(["页面", "内容"])
            for item in extract_result["text_data"]:
                ws.append([item["page"], item["content"]])

        # 保存文件
        wb.save(str(self.output_path))

        return {
            "success": True,
            "output_file": str(self.output_path),
            "method_used": method,
            "rows_written": ws.max_row if ws.max_row else 0,
        }
