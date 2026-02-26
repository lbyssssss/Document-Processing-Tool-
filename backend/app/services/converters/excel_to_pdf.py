# Excel to PDF Conversion
from openpyxl import load_workbook
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from typing import List, Optional
import io


class ExcelToPDFConverter:
    """Excel转PDF转换器"""

    def __init__(self, excel_path: Path, output_path: Path):
        self.excel_path = excel_path
        self.output_path = output_path

    def extract_content(self) -> dict:
        """从Excel文档中提取内容"""
        try:
            wb = load_workbook(str(self.excel_path))
            ws = wb.active

            # 提取数据
            data = []
            for row in ws.iter_rows(values_only=True):
                # 过滤空行
                if any(cell is not None for cell in row):
                    data.append([str(cell) if cell is not None else "" for cell in row])

            # 获取列数
            max_cols = max(len(row) for row in data) if data else 0

            return {
                "success": True,
                "data": data,
                "rows": len(data),
                "columns": max_cols,
                "sheet_name": ws.title,
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
            }

    def convert(self) -> dict:
        """将Excel转换为PDF格式"""
        try:
            # 提取Excel内容
            content = self.extract_content()
            if not content.get("success"):
                return content

            # 创建PDF文档
            pdf = SimpleDocTemplate(str(self.output_path), pagesize=A4)

            # 定义表格样式
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), colors.black),
                ('ALIGN', (0, 0), 'LEFT'),
                ('FONTNAME', (0, 0), 'Helvetica'),
                ('FONTSIZE', (0, 0), 10),
                ('BOTTOMPADDING', (0, 0), 12),
                ('BACKGROUND', (1, 0), colors.beige),
                ('GRID', (0, 1), 1, colors.black),
            ])

            if content["data"]:
                # 创建表格
                t = Table(content["data"])
                t.setStyle(table_style)
                t.wrapOn(pdf, 6.3 * inch, 8.5 * inch)
                pdf.build([t])

            return {
                "success": True,
                "output_file": str(self.output_path),
                "rows_converted": content["rows"],
                "columns_converted": content["columns"],
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
            }
