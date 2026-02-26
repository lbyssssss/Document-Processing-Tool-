# Conversion Service
from pathlib import Path
from typing import Optional, List
from datetime import datetime
import uuid
import os
import logging

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from app.models.document import DocumentType
from app.services.converters import (
    PDFToWordConverter,
    WordToPDFConverter,
    PDFToExcelConverter,
    ExcelToPDFConverter,
    PDFToPPTConverter,
    PPTToPDFConverter,
    PDFToImagesConverter,
    ImagesToPDFConverter,
)


class ConversionOptions:
    """转换选项"""

    def __init__(
        self,
        preserve_formatting: bool = True,
        quality: Optional[int] = None,
        password: Optional[str] = None,
        include_annotations: bool = False,
        dpi: Optional[int] = None,
        page_size: Optional[str] = None,
        orientation: Optional[str] = None,
    ):
        self.preserve_formatting = preserve_formatting
        self.quality = quality
        self.password = password
        self.include_annotations = include_annotations
        self.dpi = dpi
        self.page_size = page_size
        self.orientation = orientation


class ConversionResult:
    """转换结果"""

    def __init__(
        self,
        success: bool,
        output_path: Optional[str] = None,
        output_format: str = "",
        warnings: Optional[List[str]] = None,
        error: Optional[str] = None,
    ):
        self.success = success
        self.output_path = output_path
        self.output_format = output_format
        self.warnings = warnings or []
        self.error = error


class ConversionService:
    """文档转换服务"""

    def __init__(self, upload_dir: str, output_dir: str):
        self.upload_dir = Path(upload_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_output_path(self, input_filename: str, output_format: str) -> Path:
        """生成输出文件路径"""
        base_name = Path(input_filename).stem
        return self.output_dir / f"{base_name}_converted.{output_format}"

    async def pdf_to_word(
        self, file: Path, options: ConversionOptions
    ) -> ConversionResult:
        """PDF转Word"""
        try:
            logger.info(f"开始 PDF 到 Word 转换: {file.name}")
            output_path = self._get_output_path(file.name, "docx")
            logger.info(f"输出路径: {output_path}")

            converter = PDFToWordConverter(file, output_path)
            result = converter.convert()

            logger.info(f"转换器结果: {result}")

            if not result.get("success"):
                logger.error(f"转换失败: {result.get('error')}")
                return ConversionResult(
                    success=False,
                    output_format="docx",
                    error=result.get("error", "Unknown error"),
                )

            logger.info(f"转换成功，输出文件: {result.get('output_file')}")

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="docx",
                warnings=["注意：完整的PDF到Word转换需要额外的库处理"],
            )

        except Exception as e:
            logger.error(f"PDF 到 Word 转换异常: {e}")
            return ConversionResult(
                success=False,
                output_format="docx",
                error=str(e),
            )

    async def word_to_pdf(
        self, file: Path, options: ConversionOptions
    ) -> ConversionResult:
        """Word转PDF"""
        try:
            output_path = self._get_output_path(file.name, "pdf")
            converter = WordToPDFConverter(file, output_path)
            result = converter.convert()

            if not result.get("success"):
                return ConversionResult(
                    success=False,
                    output_format="pdf",
                    error=result.get("error", "Unknown error"),
                )

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="pdf",
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                output_format="pdf",
                error=str(e),
            )

    async def pdf_to_excel(
        self, file: Path, options: ConversionOptions
    ) -> ConversionResult:
        """PDF转Excel"""
        try:
            output_path = self._get_output_path(file.name, "xlsx")
            converter = PDFToExcelConverter(file, output_path)
            result = converter.convert(method="auto")

            if not result.get("success"):
                return ConversionResult(
                    success=False,
                    output_format="xlsx",
                    error=result.get("error", "Unknown error"),
                )

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="xlsx",
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                output_format="xlsx",
                error=str(e),
            )

    async def excel_to_pdf(
        self, file: Path, options: ConversionOptions
    ) -> ConversionResult:
        """Excel转PDF"""
        try:
            output_path = self._get_output_path(file.name, "pdf")
            converter = ExcelToPDFConverter(file, output_path)
            result = converter.convert()

            if not result.get("success"):
                return ConversionResult(
                    success=False,
                    output_format="pdf",
                    error=result.get("error", "Unknown error"),
                )

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="pdf",
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                output_format="pdf",
                error=str(e),
            )

    async def pdf_to_ppt(
        self, file: Path, options: ConversionOptions
    ) -> ConversionResult:
        """PDF转PPT"""
        try:
            output_path = self._get_output_path(file.name, "pptx")
            converter = PDFToPPTConverter(file, output_path)
            result = converter.convert()

            if not result.get("success"):
                return ConversionResult(
                    success=False,
                    output_format="pptx",
                    error=result.get("error", "Unknown error"),
                )

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="pptx",
                warnings=["注意：完整的PDF到PPT转换需要pdf2image库以实现图像嵌入"],
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                output_format="pptx",
                error=str(e),
            )

    async def ppt_to_pdf(
        self, file: Path, options: ConversionOptions
    ) -> ConversionResult:
        """PPT转PDF"""
        try:
            output_path = self._get_output_path(file.name, "pdf")
            converter = PPTToPDFConverter(file, output_path)
            result = converter.convert()

            if not result.get("success"):
                return ConversionResult(
                    success=False,
                    output_format="pdf",
                    error=result.get("error", "Unknown error"),
                )

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="pdf",
                warnings=["注意：完整的PPT到PDF转换需要LibreOffice或Microsoft Office以保留完整格式"],
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                output_format="pdf",
                error=str(e),
            )

    async def images_to_pdf(
        self, files: List[Path], options: ConversionOptions
    ) -> ConversionResult:
        """图片转PDF"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.output_dir / f"images_{timestamp}.pdf"

            page_size = options.page_size or "a4"
            orientation = options.orientation or "keep-original"

            converter = ImagesToPDFConverter(files, output_path, page_size, orientation)
            result = converter.convert()

            if not result.get("success"):
                return ConversionResult(
                    success=False,
                    output_format="pdf",
                    error=result.get("error", "Unknown error"),
                )

            return ConversionResult(
                success=True,
                output_path=str(output_path),
                output_format="pdf",
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                output_format="pdf",
                error=str(e),
            )

    async def pdf_to_images(
        self, file: Path, options: ConversionOptions
    ) -> List[ConversionResult]:
        """PDF转图片"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = self.output_dir / f"{file.stem}_{timestamp}"
            output_dir.mkdir(parents=True, exist_ok=True)

            dpi = options.dpi or 200
            output_format = options.quality or "png"  # 使用quality参数作为格式

            converter = PDFToImagesConverter(file, output_dir, output_format, dpi)
            result = converter.convert()

            if not result.get("success"):
                error_msg = result.get("error", "Unknown error")
                return [
                    ConversionResult(success=False, error=error_msg)
                ]

            # 返回每个图片的结果
            results = []
            for img_info in result.get("images", []):
                img_path = Path(result["output_dir"]) / img_info["filename"]
                results.append(ConversionResult(
                    success=True,
                    output_path=str(img_path),
                    output_format=output_format,
                ))

            return results

        except Exception as e:
            return [
                ConversionResult(success=False, error=str(e))
            ]

    async def get_document_info(self, file: Path) -> dict:
        """获取文档信息（用于预览）"""
        try:
            suffix = file.suffix.lower()

            if suffix == ".pdf":
                from PyPDF2 import PdfReader
                reader = PdfReader(str(file))
                return {
                    "type": "pdf",
                    "page_count": len(reader.pages),
                    "metadata": reader.metadata,
                }
            elif suffix in [".doc", ".docx"]:
                return {
                    "type": "word",
                    "page_count": None,
                    "metadata": None,
                }
            elif suffix in [".xls", ".xlsx"]:
                from openpyxl import load_workbook
                wb = load_workbook(str(file))
                ws = wb.active
                return {
                    "type": "excel",
                    "page_count": ws.max_row,
                    "metadata": None,
                }
            else:
                return {
                    "type": "unknown",
                    "page_count": None,
                    "metadata": None,
                }

        except Exception as e:
            return {
                "error": str(e),
            }
